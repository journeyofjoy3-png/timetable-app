import streamlit as st
import pandas as pd
from io import BytesIO, StringIO
import datetime
import os
import sqlite3
import hashlib
import base64
import plotly.express as px
from streamlit_option_menu import option_menu
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils.dataframe import dataframe_to_rows
import google.generativeai as genai
from PIL import Image

# --- Password Hashing Helper ---
def make_hashes(password): return hashlib.sha256(str.encode(password)).hexdigest()

def check_hashes(password, hashed_text):
    if make_hashes(password) == hashed_text: return hashed_text
    return False

# --- Database Setup ---
def init_db():
    conn = sqlite3.connect('faculty_history.db')
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS workload_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT, upload_date TEXT, week_label TEXT,
            teacher TEXT, class_lectures REAL, doubt_slots REAL, total_workload REAL,
            effective_capacity REAL, leaves REAL)''')
    c.execute('''CREATE TABLE IF NOT EXISTS users (username TEXT PRIMARY KEY, password TEXT, role TEXT)''')
    c.execute("SELECT COUNT(*) FROM users")
    if c.fetchone()[0] == 0:
        c.execute("INSERT INTO users (username, password, role) VALUES (?, ?, ?)", ("admin", make_hashes("matrix2026"), "admin"))
    conn.commit()
    conn.close()

init_db()

def login_user(username, password):
    conn = sqlite3.connect('faculty_history.db')
    c = conn.cursor()
    c.execute('SELECT * FROM users WHERE username =? AND password = ?', (username, make_hashes(password)))
    data = c.fetchall()
    conn.close()
    return data

def add_user(username, password, role="admin"):
    conn = sqlite3.connect('faculty_history.db')
    c = conn.cursor()
    try:
        c.execute('INSERT INTO users(username, password, role) VALUES (?,?,?)', (username, make_hashes(password), role))
        conn.commit()
        conn.close()
        return True
    except sqlite3.IntegrityError:
        conn.close()
        return False

def update_password(username, new_password):
    conn = sqlite3.connect('faculty_history.db')
    c = conn.cursor()
    c.execute('UPDATE users SET password = ? WHERE username = ?', (make_hashes(new_password), username))
    conn.commit()
    conn.close()

def save_to_db(df, week_label):
    conn = sqlite3.connect('faculty_history.db')
    c = conn.cursor()
    c.execute("DELETE FROM workload_history WHERE week_label = ?", (week_label,))
    df_to_save = df[['Teacher', 'Class Lectures', 'Doubt Slots', 'Total Workload', 'Effective Capacity', 'Leave Days Count']].copy()
    df_to_save['week_label'] = week_label
    df_to_save['upload_date'] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    df_to_save.rename(columns={'Teacher': 'teacher', 'Class Lectures': 'class_lectures', 'Doubt Slots': 'doubt_slots', 'Total Workload': 'total_workload', 'Effective Capacity': 'effective_capacity', 'Leave Days Count': 'leaves'}, inplace=True)
    df_to_save.to_sql('workload_history', conn, if_exists='append', index=False)
    conn.close()

# --- AI Image Parsing Logic ---
def extract_timetable_from_image(image_file, api_key):
    genai.configure(api_key=api_key)
    # Using the fast & highly capable 1.5-flash model
    model = genai.GenerativeModel('gemini-1.5-flash') 
    img = Image.open(image_file)
    
    prompt = """
    You are an expert data extraction AI. I am providing an image of a faculty timetable.
    Extract this timetable into a structured CSV format.
    The columns MUST EXACTLY match this comma-separated header:
    Teacher,Monday_M1,Monday_M2,Monday_M3,Monday_M4,Monday_E1,Monday_E2,Monday_E3,Monday_E4,Tuesday_M1,Tuesday_M2,Tuesday_M3,Tuesday_M4,Tuesday_E1,Tuesday_E2,Tuesday_E3,Tuesday_E4,Wednesday_M1,Wednesday_M2,Wednesday_M3,Wednesday_M4,Wednesday_E1,Wednesday_E2,Wednesday_E3,Wednesday_E4,Thursday_M1,Thursday_M2,Thursday_M3,Thursday_M4,Thursday_E1,Thursday_E2,Thursday_E3,Thursday_E4,Friday_M1,Friday_M2,Friday_M3,Friday_M4,Friday_E1,Friday_E2,Friday_E3,Friday_E4,Saturday_M1,Saturday_M2,Saturday_M3,Saturday_M4,Saturday_E1,Saturday_E2,Saturday_E3,Saturday_E4
    
    Rules:
    1. Read each row (each teacher). 
    2. If a cell in the image has a class name/batch, put that text in the corresponding CSV column.
    3. If a cell is empty or blank, leave the CSV value empty.
    4. Do NOT use markdown code blocks like ```csv. Output raw CSV text only.
    """
    response = model.generate_content([prompt, img])
    
    # Safely strip markdown backticks to prevent syntax errors
    raw_text = response.text
    raw_text = raw_text.replace('```csv', '')
    raw_text = raw_text.replace('```', '')
    csv_data = raw_text.strip()
    
    df = pd.read_csv(StringIO(csv_data))
    return df

# --- Page Setup & CSS ---
st.set_page_config(page_title="Matrix Net - Faculty Portal", layout="wide", page_icon="🏫")
st.markdown("""
<style>
    div[data-testid="metric-container"] { background: linear-gradient(135deg, #1f2937 0%, #111827 100%); border: 1px solid #374151; border-radius: 12px; padding: 16px 20px; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.2); transition: transform 0.2s ease, box-shadow 0.2s ease; }
    div[data-testid="metric-container"]:hover { transform: translateY(-2px); box-shadow: 0 10px 15px -3px rgba(0,0,0,0.3); }
    div[data-testid="stMetricValue"] { font-size: 28px !important; font-weight: 700 !important; color: #38bdf8 !important; }
    div[data-testid="stMetricLabel"] { font-size: 14px !important; font-weight: 600 !important; color: #9ca3af !important; }
    .stButton > button { border-radius: 8px !important; font-weight: 600 !important; transition: all 0.2s ease-in-out !important; }
    .ai-question-box { background-color: #1e293b; padding: 25px; border-radius: 12px; border-left: 5px solid #38bdf8; margin-top: 20px;}
</style>
""", unsafe_allow_html=True)

# --- Authentication & Lockscreen ---
if "authenticated" not in st.session_state: st.session_state["authenticated"] = False
if "username" not in st.session_state: st.session_state["username"] = ""

if not st.session_state["authenticated"]:
    def set_login_background(image_file):
        if os.path.exists(image_file):
            with open(image_file, "rb") as f:
                encoded_string = base64.b64encode(f.read()).decode()
            css = f"""
            <style>
            .stApp {{ background-image: url(data:image/jpeg;base64,{encoded_string}); background-size: cover; background-position: center; background-repeat: no-repeat; background-attachment: fixed; }}
            .stApp::before {{ content: ""; position: absolute; top: 0; left: 0; width: 100%; height: 100%; background-color: rgba(15, 23, 42, 0.6); z-index: -1; }}
            [data-testid="column"]:nth-of-type(2) {{ background: rgba(30, 41, 59, 0.7); backdrop-filter: blur(12px); -webkit-backdrop-filter: blur(12px); padding: 40px; border-radius: 16px; border: 1px solid rgba(255, 255, 255, 0.1); box-shadow: 0 25px 50px -12px rgba(0,0,0,0.5); }}
            </style>
            """
            st.markdown(css, unsafe_allow_html=True)

    set_login_background("background.jpg")
    col1, col2, col3 = st.columns([1, 1.5, 1]) 
    with col2:
        if os.path.exists("logo.png"):
            st.markdown("""<div style="text-align: center;">""", unsafe_allow_html=True)
            st.image("logo.png", width=200)
            st.markdown("</div>", unsafe_allow_html=True)
        else:
            st.markdown("<h2 style='text-align: center; color: #38bdf8;'>🟦 MATRIX NET</h2>", unsafe_allow_html=True)
        st.markdown("<h3 style='text-align: center;'>🔒 Secure Portal</h3>", unsafe_allow_html=True)
        st.write("")
        with st.form("login_form"):
            username = st.text_input("User ID")
            password = st.text_input("Password", type="password")
            submitted = st.form_submit_button("Login", type="primary", use_container_width=True)
            if submitted:
                result = login_user(username, password)
                if result:
                    st.session_state["authenticated"], st.session_state["username"] = True, username
                    st.rerun()
                else:
                    st.error("Incorrect User ID or Password.")
    st.stop()

# --- Sidebar Navigation ---
with st.sidebar:
    if os.path.exists("logo.png"): st.image("logo.png", use_container_width=True)
    else: st.markdown("### 🟦 MATRIX NET")
    st.caption(f"Logged in as: **{st.session_state['username']}**")
    st.divider()
    
    st.markdown("### ⚙️ AI Settings")
    ai_api_key = st.text_input("Gemini API Key", type="password", help="Required for Image uploads and AI Question Generation.")
    st.divider()
    
    selected = option_menu(
        "Main Menu", 
        ["Timetable Analyzer", "Schedule Calculator", "Doubt Generator", "Teacher Dashboard", "Historical Analytics", "Question Generator", "User Management"], 
        icons=["bar-chart-fill", "calendar3", "robot", "person-badge", "graph-up", "patch-question", "gear-fill"], 
        menu_icon="cast", 
        default_index=0, 
        styles={"nav-link-selected": {"background-color": "#38bdf8", "color": "white"}}
    )
    st.divider()
    if st.button("🚪 Logout", use_container_width=True):
        st.session_state["authenticated"], st.session_state["username"] = False, ""
        st.rerun()

# ==========================================
# PAGE 1: TIMETABLE ANALYZER
# ==========================================
if selected == "Timetable Analyzer":
    st.header("📊 Faculty Workload Analyzer")
    with st.expander("📁 Upload / Update Timetables", expanded=True):
        mode = st.radio("Select Analysis Mode:", ["Class Timetable Only", "Class + Doubt Timetables"], horizontal=True)
        class_file, doubt_file = None, None
        if mode == "Class Timetable Only":
            class_file = st.file_uploader("Upload Weekly Class Timetable (Excel or Image)", type=["xlsx", "xls", "png", "jpg", "jpeg"], key="class_only")
        else:
            upload_col1, upload_col2 = st.columns(2)
            with upload_col1: class_file = st.file_uploader("1. Upload Class Timetable (Excel or Image)", type=["xlsx", "xls", "png", "jpg", "jpeg"], key="class_file")
            with upload_col2: doubt_file = st.file_uploader("2. Upload Doubt Timetable (Excel ONLY)", type=["xlsx", "xls"], key="doubt_file")

    def parse_class_timetable(df, is_ai_format=False):
        days_map = {"Monday": range(1,9), "Tuesday": range(9,17), "Wednesday": range(17,25), "Thursday": range(26,34), "Friday": range(34,42), "Saturday": range(42,50)}
        days, slots = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"], ["M1", "M2", "M3", "M4", "E1", "E2", "E3", "E4"]
        records, teacher_stats = [], {}
        total_raw_assigned_slots, total_raw_teachers = 0, 0
        
        if is_ai_format:
            for index, row in df.iterrows():
                teacher_name = str(row.get('Teacher', '')).strip()
                if not teacher_name or teacher_name.lower() == 'nan': continue
                total_raw_teachers += 1
                valid_classes, total_classes, day_counts = {}, 0, {day: 0 for day in days}
                
                for day in days:
                    for slot in slots:
                        col_name = f"{day}_{slot}"
                        if col_name in df.columns:
                            val = str(row[col_name]).strip()
                            if val and val.lower() != 'nan':
                                total_raw_assigned_slots += 1; total_classes += 1; day_counts[day] += 1
                                if (val, val) not in valid_classes: valid_classes[(val, val)] = 0
                                valid_classes[(val, val)] += 1
                
                leave_days_list = [day for day, cnt in day_counts.items() if cnt == 0] if total_classes > 0 else []
                status, effective_capacity = ("Active", 48 - (len(leave_days_list) * 8)) if total_classes > 0 else ("Not Allotted", 48)
                leave_text = ", ".join(leave_days_list) if leave_days_list else ("None" if status == "Active" else "N/A")
                teacher_stats[teacher_name] = {'Total_Classes': total_classes, 'Effective_Capacity': effective_capacity, 'Leave_Count': len(leave_days_list), 'Status': status, 'Leave_Days': leave_text}
                if not valid_classes: records.append({'Teacher': teacher_name, 'Subject': 'None', 'Batch': 'None', 'Lectures': 0})
                else:
                    for (subj, batch), count in valid_classes.items(): records.append({'Teacher': teacher_name, 'Subject': subj, 'Batch': batch, 'Lectures': count})
        else:
            i = 2
            while i < len(df):
                teacher_name = str(df.iloc[i, 0]).strip()
                if teacher_name == '' or teacher_name.lower() == 'nan': i += 1; continue
                total_raw_teachers += 1
                subject_row = df.iloc[i, :]
                has_batch_row = False
                if i + 1 < len(df) and (str(df.iloc[i+1, 0]).strip() == '' or str(df.iloc[i+1, 0]).strip().lower() == 'nan'):
                    has_batch_row, batch_row = True, df.iloc[i+1, :]
                else: batch_row = pd.Series([''] * len(subject_row))
                valid_classes, total_classes, day_counts = {}, 0, {day: 0 for day in days_map}
                for day, cols in days_map.items():
                    for col in cols:
                        if col >= len(subject_row): continue
                        subj, batch = str(subject_row.iloc[col]).strip(), str(batch_row.iloc[col]).strip()
                        if (subj != '' and subj.lower() != 'nan') and (batch != '' and batch.lower() != 'nan'):
                            total_raw_assigned_slots, total_classes, day_counts[day] = total_raw_assigned_slots+1, total_classes+1, day_counts[day]+1
                            if (subj, batch) not in valid_classes: valid_classes[(subj, batch)] = 0
                            valid_classes[(subj, batch)] += 1
                leave_days_list = [day for day, cnt in day_counts.items() if cnt == 0] if total_classes > 0 else []
                status, effective_capacity = ("Active", 48 - (len(leave_days_list) * 8)) if total_classes > 0 else ("Not Allotted", 48)
                leave_text = ", ".join(leave_days_list) if leave_days_list else ("None" if status == "Active" else "N/A")
                teacher_stats[teacher_name] = {'Total_Classes': total_classes, 'Effective_Capacity': effective_capacity, 'Leave_Count': len(leave_days_list), 'Status': status, 'Leave_Days': leave_text}
                if not valid_classes: records.append({'Teacher': teacher_name, 'Subject': 'None', 'Batch': 'None', 'Lectures': 0})
                else:
                    for (subj, batch), count in valid_classes.items(): records.append({'Teacher': teacher_name, 'Subject': subj, 'Batch': batch, 'Lectures': count})
                i += 2 if has_batch_row else 1
        return pd.DataFrame(records), teacher_stats, total_raw_teachers, total_raw_assigned_slots

    def parse_doubt_timetable(uploaded_file):
        df = pd.read_excel(uploaded_file, sheet_name=0, keep_default_na=False)
        doubt_counts, total_doubt_slots = {}, 0
        for index, row in df.iterrows():
            teacher_name = str(row.iloc[0]).strip()
            if teacher_name == 'nan' or not teacher_name or teacher_name == "Teacher's Name": continue
            slots_count = 0
            for day in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']:
                if day in df.columns:
                    slots_str = str(row[day]).strip()
                    if slots_str != 'nan' and slots_str: slots_count += len([s for s in slots_str.split(',') if s.strip()])
            doubt_counts[teacher_name] = slots_count
            total_doubt_slots += slots_count
        return doubt_counts, total_doubt_slots

    if class_file is not None:
        is_image = class_file.name.lower().endswith(('png', 'jpg', 'jpeg'))
        if is_image and not ai_api_key:
            st.error("🔑 Please enter your Gemini API Key in the sidebar to process images.")
        elif mode == "Class + Doubt Timetables" and doubt_file is None: 
            st.warning("Please upload the Doubt Timetable to proceed.")
        else:
            with st.spinner("🤖 Processing Timetables..." if not is_image else "🤖 AI is reading the image... This may take a minute."):
                if is_image:
                    df_raw_class = extract_timetable_from_image(class_file, ai_api_key)
                    df_tidy, class_teacher_stats, total_teachers, total_class_slots = parse_class_timetable(df_raw_class, is_ai_format=True)
                else:
                    df_raw_class = pd.read_excel(class_file, sheet_name=0, keep_default_na=False)
                    df_tidy, class_teacher_stats, total_teachers, total_class_slots = parse_class_timetable(df_raw_class, is_ai_format=False)
                
                doubt_slots_map, total_doubt_slots = {}, 0
                if mode == "Class + Doubt Timetables" and doubt_file is not None: doubt_slots_map, total_doubt_slots = parse_doubt_timetable(doubt_file)
                summary_rows = []
                for teacher, info in class_teacher_stats.items():
                    class_lecs, doubt_lecs, effective_capacity = info['Total_Classes'], doubt_slots_map.get(teacher, 0), info['Effective_Capacity']
                    total_workload, final_free_slots = class_lecs + doubt_lecs, max(0, effective_capacity - (class_lecs + doubt_lecs))
                    summary_rows.append({'Teacher': teacher, 'Class Lectures': class_lecs, 'Doubt Slots': doubt_lecs, 'Total Workload': total_workload, 'Leave / Off Days': info['Leave_Days'], 'Leave Days Count': info['Leave_Count'], 'Effective Capacity': effective_capacity, 'Net Free Slots': final_free_slots, 'True Class Util.': class_lecs / effective_capacity if effective_capacity > 0 else 0, 'True Total Util.': total_workload / effective_capacity if effective_capacity > 0 else 0, 'Status': info['Status']})
                df_summary = pd.DataFrame(summary_rows).sort_values(by='Total Workload', ascending=False)
                st.session_state["df_summary_cached"] = df_summary
                
            st.markdown("##### 📈 Key Workload Overview")
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Total Faculty", total_teachers)
            m2.metric("Class Lectures", total_class_slots)
            m3.metric("Doubt Slots", total_doubt_slots if mode == "Class + Doubt Timetables" else "N/A")
            m4.metric("Leaves Taken", df_summary['Leave Days Count'].sum())
            
            st.divider()
            st.markdown("##### 💾 Save to Historical Database")
            sc1, sc2 = st.columns([3, 1])
            with sc1: week_label = st.text_input("Timetable Label (e.g., 'Week of Aug 10')", placeholder="Week of Aug 10 - Aug 15")
            with sc2:
                st.write(""); st.write("")
                if st.button("Save to Archive", use_container_width=True):
                    if week_label:
                        save_to_db(df_summary, week_label)
                        st.success(f"Data saved to database under '{week_label}'!"); st.toast("Saved to History!", icon="💾")
                    else: st.error("Please provide a week label to save.")
                        
            st.divider()
            st.markdown("##### 📊 Workload Distribution (Top 15 Busiest)")
            fig = px.bar(df_summary.head(15), x='Teacher', y=['Class Lectures', 'Doubt Slots'], barmode='stack', color_discrete_sequence=['#38bdf8', '#8b5cf6'])
            fig.update_layout(plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", font_color="#f8fafc", margin=dict(t=10, b=10))
            st.plotly_chart(fig, use_container_width=True)
            
            st.markdown("##### 📋 Faculty Workload Summary Table")
            display_df = df_summary.drop(columns=['Leave Days Count', 'Doubt Slots', 'True Total Util.', 'Total Workload']) if mode == "Class Timetable Only" else df_summary.drop(columns=['Leave Days Count'])
            st.dataframe(display_df.style.format({'True Class Util.': '{:.1%}', 'True Total Util.': '{:.1%}'}), use_container_width=True)

# ==========================================
# PAGE 2: SCHEDULE CALCULATOR 
# ==========================================
elif selected == "Schedule Calculator":
    st.header("📅 Syllabus Timeline Calculator")
    calc_option = st.selectbox("Calculate:", ["End Date", "Total Number of Lectures", "Lectures per Week", "Start Date"])
    col_a, col_b = st.columns(2)
    if calc_option == "End Date":
        with col_a:
            start_date, total_lec = st.date_input("Start Date"), st.number_input("Total Number of Lectures", min_value=1, value=120)
        with col_b:
            lec_per_week, leave_days = st.number_input("Lectures per Week", min_value=0.5, value=6.0, step=0.5), st.number_input("Planned Holidays / Off Days", min_value=0, value=0, step=1)
        if st.button("Calculate", type="primary"): st.success(f"### 🎯 Course Completion Date: **{(start_date + datetime.timedelta(days=((total_lec / lec_per_week) * 7) + leave_days)).strftime('%B %d, %Y')}**")

    elif calc_option == "Total Number of Lectures":
        with col_a:
            start_date, end_date = st.date_input("Start Date"), st.date_input("End Date", value=datetime.date.today() + datetime.timedelta(days=90))
        with col_b:
            lec_per_week, leave_days = st.number_input("Lectures per Week", min_value=0.5, value=6.0, step=0.5), st.number_input("Planned Holidays", min_value=0, value=0, step=1)
        if st.button("Calculate", type="primary"):
            effective_days = (end_date - start_date).days - leave_days
            if effective_days > 0: st.success(f"### 🎯 Total Possible Lectures: **{int((effective_days / 7) * lec_per_week)}**")

    elif calc_option == "Lectures per Week":
        with col_a:
            start_date, end_date = st.date_input("Start Date"), st.date_input("End Date", value=datetime.date.today() + datetime.timedelta(days=90))
        with col_b:
            total_lec, leave_days = st.number_input("Total Required Lectures", min_value=1, value=120), st.number_input("Planned Holidays", min_value=0, value=0, step=1)
        if st.button("Calculate", type="primary"):
            effective_days = (end_date - start_date).days - leave_days
            if effective_days > 0: st.success(f"### 🎯 Required Weekly Pace: **{total_lec / (effective_days / 7):.1f}** lec/week")

    elif calc_option == "Start Date":
        with col_a:
            end_date, total_lec = st.date_input("Target End Date"), st.number_input("Total Required Lectures", min_value=1, value=120)
        with col_b:
            lec_per_week, leave_days = st.number_input("Lectures per Week", min_value=0.5, value=6.0, step=0.5), st.number_input("Planned Holidays", min_value=0, value=0, step=1)
        if st.button("Calculate", type="primary"): st.success(f"### 🎯 Required Course Start Date: **{(end_date - datetime.timedelta(days=((total_lec / lec_per_week) * 7) + leave_days)).strftime('%B %d, %Y')}**")

# ==========================================
# PAGE 3: DOUBT GENERATOR (WITH SUBJECTS)
# ==========================================
elif selected == "Doubt Generator":
    st.header("🤖 Auto Doubt Generator")
    with st.expander("⚙️ Upload & Config", expanded=True):
        generator_class_file = st.file_uploader("1. Upload Class Timetable (Excel or Image)", type=["xlsx", "xls", "png", "jpg", "jpeg"], key="gen_class")
        input_method = st.radio("How should we select teachers for doubt slots?", ["Auto-detect all faculty from timetable", "Upload Subject Mapping Excel (Teacher & Subject columns)", "Enter specific faculty names manually"])
        mapping_file, teacher_input, subject_name = None, "", ""
        
        if input_method == "Upload Subject Mapping Excel (Teacher & Subject columns)":
            st.info("Upload an Excel file with a 'Teacher' column and a 'Subject' column. The final output will be beautifully separated into subject tabs!")
            mapping_file = st.file_uploader("2. Upload Master Mapping File (Excel)", type=["xlsx", "xls"], key="map_file")
        elif input_method == "Enter specific faculty names manually":
            subject_name = st.text_input("Enter Subject Name (Optional, e.g., PHYSICS, CHEMISTRY)")
            teacher_input = st.text_area("Enter Faculty Names (Comma Separated)", height=100)
    
    if st.button("✨ Generate Schedule", type="primary"):
        is_image = generator_class_file is not None and generator_class_file.name.lower().endswith(('png', 'jpg', 'jpeg'))
        
        if generator_class_file is None: st.error("Please upload the Class Timetable file.")
        elif is_image and not ai_api_key: st.error("🔑 Please enter your Gemini API Key in the sidebar to process images.")
        elif input_method == "Enter specific faculty names manually" and not teacher_input.strip(): st.error("Please enter at least one faculty code.")
        elif input_method == "Upload Subject Mapping Excel (Teacher & Subject columns)" and mapping_file is None: st.error("Please upload the Teacher-Subject mapping Excel file.")
        else:
            with st.spinner("🤖 AI is reading the image..." if is_image else "Crunching the timetable..."):
                if is_image:
                    df_class = extract_timetable_from_image(generator_class_file, ai_api_key)
                    is_ai_format = True
                else:
                    df_class = pd.read_excel(generator_class_file, sheet_name=0, keep_default_na=False)
                    is_ai_format = False
                
                days_map = {"Monday": 1, "Tuesday": 9, "Wednesday": 17, "Thursday": 26, "Friday": 34, "Saturday": 42}
                
                def has_class(row, day, start_col, slot_idx, is_ai):
                    if is_ai:
                        slots = ["M1", "M2", "M3", "M4", "E1", "E2", "E3", "E4"]
                        col_name = f"{day}_{slots[slot_idx]}"
                        if col_name in row.index:
                            val = str(row[col_name]).strip()
                            return (val != '' and val.lower() != 'nan')
                        return False
                    else:
                        col = start_col + slot_idx
                        if col >= len(row): return False
                        val = str(row.iloc[col]).strip()
                        return (val != '' and val.lower() != 'nan')

                subject_mapping = {}
                
                if input_method == "Auto-detect all faculty from timetable":
                    if is_ai_format:
                        raw_teachers = df_class['Teacher'].dropna().astype(str).str.strip().unique()
                    else:
                        raw_teachers = df_class.iloc[2:, 0].dropna().astype(str).str.strip().unique()
                    target_teachers = [t for t in raw_teachers if t and t.lower() not in ['nan', 'teacher', "teacher's name"]]
                    for t in target_teachers: subject_mapping[t] = "All Faculty"
                elif input_method == "Upload Subject Mapping Excel (Teacher & Subject columns)":
                    df_map = pd.read_excel(mapping_file)
                    target_teachers = []
                    if 'Teacher' in df_map.columns and 'Subject' in df_map.columns:
                        for index, row in df_map.iterrows():
                            t, s = str(row['Teacher']).strip(), str(row['Subject']).strip()
                            if t and t.lower() != 'nan':
                                target_teachers.append(t)
                                subject_mapping[t] = s.upper() if s and s.lower() != 'nan' else "Unassigned"
                    else:
                        st.error("Mapping file must have exact column headers 'Teacher' and 'Subject'.")
                        st.stop()
                else:
                    target_teachers = [t.strip() for t in teacher_input.split(",") if t.strip()]
                    for t in target_teachers: subject_mapping[t] = subject_name.upper() if subject_name else "All Faculty"

                if not target_teachers:
                    st.error("No valid faculty names were found. Please check your file or input.")
                else:
                    generated_doubts = []

                    for teacher in target_teachers:
                        if is_ai_format:
                            teacher_idx = df_class[df_class['Teacher'] == teacher].index
                        else:
                            teacher_idx = df_class[df_class.iloc[:, 0] == teacher].index
                            
                        teacher_schedule = {"Teacher": teacher, "Subject": subject_mapping.get(teacher, "Unassigned")}
                        row = df_class.iloc[teacher_idx[0], :] if not teacher_idx.empty else pd.Series([''] * 50)

                        for day, start_col in days_map.items():
                            c0, c1, c2, c3, c4, c5, c6, c7 = [has_class(row, day, start_col, i, is_ai_format) for i in range(8)]
                            class_count = sum([c0, c1, c2, c3, c4, c5, c6, c7])
                            
                            d1_avail, d2_avail, d3_avail = not c2, not (c4 and c5 and c6), not c6  
                            doubts_needed = max(0, 5 - class_count) if class_count > 0 else 3
                            
                            preference = []
                            if d2_avail: preference.append("D2")
                            if d3_avail: preference.append("D3")
                            if d1_avail: preference.append("D1")
                            
                            assigned_doubts = sorted(preference[:doubts_needed])
                            teacher_schedule[day] = ", ".join(assigned_doubts) if assigned_doubts else "None"
                        generated_doubts.append(teacher_schedule)
                        
                    df_gen = pd.DataFrame(generated_doubts)
                    cols = ['Subject', 'Teacher', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
                    df_gen = df_gen[cols]
                    st.session_state["df_doubt_gen"] = df_gen
                    
                    st.success(f"✨ Process Complete! Generated schedule for {len(target_teachers)} faculty members.")
                    st.dataframe(df_gen, use_container_width=True)

                    def generate_multi_tab_excel(df):
                        wb = Workbook()
                        wb.remove(wb.active) 
                        header_font = Font(bold=True, color="FFFFFF")
                        header_fill = PatternFill("solid", fgColor="1E293B")
                        center_align = Alignment(horizontal="center", vertical="center")
                        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
                        
                        subjects = df['Subject'].unique()
                        for subj in subjects:
                            ws = wb.create_sheet(title=str(subj)[:31].replace("/","-"))
                            subj_df = df[df['Subject'] == subj].drop(columns=['Subject'])
                            for r in dataframe_to_rows(subj_df, index=False, header=True): ws.append(r)
                            for cell in ws[1]: cell.font, cell.fill, cell.alignment = header_font, header_fill, center_align
                            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                                for cell in row:
                                    cell.border = border
                                    cell.alignment = center_align
                            for col in ws.columns:
                                max_len = max([len(str(cell.value)) for cell in col if cell.value is not None] + [0])
                                ws.column_dimensions[col[0].column_letter].width = max_len + 4
                                
                        output = BytesIO()
                        wb.save(output)
                        output.seek(0)
                        return output
                        
                    excel_data = generate_multi_tab_excel(df_gen)
                    st.download_button("📥 Download Generated Doubt Timetable (Excel)", data=excel_data, file_name="MatrixNet_Doubt_Schedule.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ==========================================
# PAGE 4: TEACHER DASHBOARD
# ==========================================
elif selected == "Teacher Dashboard":
    st.header("👤 Individual Faculty Dashboard")
    if "df_summary_cached" in st.session_state:
        df_sum = st.session_state["df_summary_cached"]
        teacher_list = df_sum["Teacher"].tolist()
        selected_teacher = st.selectbox("Select Faculty Member:", teacher_list)
        if selected_teacher:
            t_data = df_sum[df_sum["Teacher"] == selected_teacher].iloc[0]
            st.markdown(f"### 📌 Profile: **{selected_teacher}**")
            sc1, sc2, sc3, sc4 = st.columns(4)
            sc1.metric("Class Lectures", t_data["Class Lectures"])
            sc2.metric("Doubt Slots", t_data.get("Doubt Slots", 0))
            sc3.metric("Effective Capacity", f"{t_data['Effective Capacity']} Slots")
            sc4.metric("Free Slots Left", t_data["Net Free Slots"])
    else:
        st.info("💡 Please upload a timetable in the Analyzer menu first.")

# ==========================================
# PAGE 5: HISTORICAL ANALYTICS
# ==========================================
elif selected == "Historical Analytics":
    st.header("📈 Historical Workload Analytics")
    conn = sqlite3.connect('faculty_history.db')
    df_hist = pd.read_sql('SELECT * FROM workload_history', conn)
    conn.close()
    
    if df_hist.empty:
        st.info("No historical data found. Please analyze a timetable and click 'Save to Archive' first.")
    else:
        teacher_list = df_hist['teacher'].unique().tolist()
        selected_teacher_hist = st.selectbox("Select Faculty to View Trends:", teacher_list)
        df_teacher_hist = df_hist[df_hist['teacher'] == selected_teacher_hist]
        
        st.markdown(f"##### Workload Trend for {selected_teacher_hist}")
        fig = px.line(df_teacher_hist, x='week_label', y=['class_lectures', 'doubt_slots', 'total_workload'], markers=True)
        fig.update_layout(plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", font_color="#f8fafc")
        st.plotly_chart(fig, use_container_width=True)
        st.dataframe(df_teacher_hist[['week_label', 'class_lectures', 'doubt_slots', 'total_workload', 'leaves']], use_container_width=True)

# ==========================================
# PAGE 6: QUESTION GENERATOR (NEW AI FEATURE)
# ==========================================
elif selected == "Question Generator":
    st.header("🧠 NEET-UG Expert Question Generator")
    st.caption("Generate highly unique, conceptual, zero-repetition MCQs using NTA expert logic.")

    col1, col2 = st.columns(2)
    with col1:
        subject = st.selectbox("Subject", ["Physics", "Chemistry", "Biology (Botany)", "Biology (Zoology)"])
        chapter = st.text_input("Chapter Name", placeholder="e.g., Thermodynamics")
        difficulty = st.selectbox("Difficulty Level", ["Easy", "Medium", "Hard", "Rank-Decider"])
    with col2:
        topic = st.text_input("Topic Name", placeholder="e.g., Carnot Engine")
        subtopic = st.text_input("Subtopic Name (Optional)", placeholder="e.g., Efficiency Graphs")

    if st.button("✨ Generate Unique MCQ", type="primary"):
        if not ai_api_key:
            st.error("🔑 Please enter your Gemini API Key in the sidebar to generate questions.")
        elif not chapter or not topic:
            st.warning("⚠️ Please fill in at least the Chapter and Topic fields.")
        else:
            with st.spinner("🤖 NTA Expert AI is crafting a unique question..."):
                genai.configure(api_key=ai_api_key)
                model = genai.GenerativeModel('gemini-1.5-flash')
                
                prompt = f"""
                Act as an expert NTA (National Testing Agency) paper setter for the NEET-UG examination. Your task is to generate a highly unique, conceptual, and zero-repetition Multiple Choice Question (MCQ) based on the parameters provided below.

                PARAMETERS:
                - Subject: {subject}
                - Chapter: {chapter}
                - Topic: {topic}
                - Subtopic: {subtopic if subtopic else 'General'}
                - Difficulty Level: {difficulty}

                STRICT ANTI-REPETITION RULES:
                To ensure this question is entirely unique and unlike standard generic questions:
                1. Concept Angle: Do not ask the most obvious definition or formula. Instead, pick a niche edge-case, an exception, a graphical interpretation, or a secondary application of the subtopic.
                2. Cross-linking: If possible, slightly integrate a foundational concept from earlier in the chapter to test comprehensive understanding.
                3. NCERT Bound: The core concept MUST strictly lie within the NCERT syllabus boundaries, but the application should require critical thinking.

                QUESTION FORMAT REQUIRED (Choose ONE randomly to ensure variety):
                - Format A: Standard Conceptual / Numerical (with a slight twist)
                - Format B: Assertion and Reason
                - Format C: Multi-Statement (e.g., "How many of the above statements are correct?")
                - Format D: Match the Columns (minimum 4 items per column)
                - Format E: "Which of the following is INCORRECT?" (Requires reading all options carefully)

                OUTPUT STRUCTURE:
                1. **Question:** (Write the question clearly. If it's physics/physical chemistry, ensure the math is solvable within 1 minute without a calculator).
                2. **Options:**
                   A) [Option 1]
                   B) [Option 2]
                   C) [Option 3]
                   D) [Option 4]
                (Ensure options include common student traps/miscalculations as distractors).
                3. **Correct Answer:** (Just the letter)
                4. **Detailed Solution & Thought Process:** (Explain WHY the right answer is right, WHERE students usually make mistakes in this specific question, and briefly explain why the distractors are wrong).
                5. **Concept Tag:** (State the specific 1-line concept or NCERT page reference tested here).
                """
                
                try:
                    response = model.generate_content(prompt)
                    st.markdown("<div class='ai-question-box'>", unsafe_allow_html=True)
                    st.markdown("### 📝 Generated Question")
                    st.markdown(response.text)
                    st.markdown("</div>", unsafe_allow_html=True)
                except Exception as e:
                    st.error(f"❌ Error generating question: {e}")

# ==========================================
# PAGE 7: USER MANAGEMENT
# ==========================================
elif selected == "User Management":
    st.header("⚙️ User Management & Security")
    tab_user1, tab_user2 = st.tabs(["🔑 Change Password", "➕ Add New User"])
    
    with tab_user1:
        st.markdown(f"##### Change Password for **{st.session_state['username']}**")
        col_p1, col_p2 = st.columns(2)
        with col_p1:
            with st.form("change_password_form"):
                current_pass = st.text_input("Current Password", type="password")
                new_pass = st.text_input("New Password", type="password")
                confirm_pass = st.text_input("Confirm New Password", type="password")
                submitted = st.form_submit_button("Update Password", type="primary")
                if submitted:
                    check_user = login_user(st.session_state['username'], current_pass)
                    if not check_user: st.error("Current password is incorrect.")
                    elif new_pass.strip() == "": st.error("New password cannot be empty.")
                    elif new_pass != confirm_pass: st.error("New passwords do not match.")
                    else:
                        update_password(st.session_state['username'], new_pass)
                        st.success("Password updated successfully!")

    with tab_user2:
        st.markdown("##### Create a New Portal Account")
        col_u1, col_u2 = st.columns(2)
        with col_u1:
            with st.form("add_new_user_form"):
                new_username = st.text_input("New User ID / Username")
                new_user_pass = st.text_input("New Account Password", type="password")
                confirm_user_pass = st.text_input("Confirm Account Password", type="password")
                submitted_new = st.form_submit_button("Create Account", type="primary")
                if submitted_new:
                    if new_username.strip() == "": st.error("Username cannot be empty.")
                    elif new_user_pass.strip() == "": st.error("Password cannot be empty.")
                    elif new_user_pass != confirm_user_pass: st.error("Passwords do not match.")
                    else:
                        success = add_user(new_username, new_user_pass)
                        if success: st.success(f"Account '{new_username}' created successfully!")
                        else: st.error(f"Username '{new_username}' already exists.")